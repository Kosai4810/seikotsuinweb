#!/usr/bin/env python3
"""
整骨院 自動営業メール送信スクリプト
- 問い合わせフォームへの自動送信
- 失敗時は手動リストへ移動
- 1日10〜20件、ランダム間隔で送信
- リスト枯渇時は新規追加
"""

import os
import sys
import json
import time
import random
import logging
from datetime import datetime, timedelta
from pathlib import Path

# 依存関係チェック
try:
    import openpyxl
    from playwright.sync_api import sync_playwright
except ImportError as e:
    print(f"依存関係エラー: {e}")
    print("pip3 install openpyxl playwright --break-system-packages")
    sys.exit(1)

# 設定
BASE_DIR = Path(__file__).parent
EXCEL_FILE = BASE_DIR / "整骨院web_営業リスト_個別営業文付き.xlsx"
LOG_FILE = BASE_DIR / "送信ログ.json"
STATE_FILE = BASE_DIR / "送信状態.json"
SCREENSHOT_DIR = BASE_DIR / "送信スクリーンショット"

# 送信者情報
SENDER_INFO = {
    "name": "WEBBRIO 柴藤",
    "name_kana": "ウェブリオ シバトウ",
    "email": "contact@seikotsuinweb.com",
    "tel": "070-7615-5231",
}

# 1日の送信上限
DAILY_MIN = 10
DAILY_MAX = 20

# 送信間隔（秒）
MIN_INTERVAL = 300   # 5分
MAX_INTERVAL = 1800  # 30分

# ログ設定
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(BASE_DIR / "auto_sender.log", encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)


def load_state():
    """送信状態を読み込む"""
    if STATE_FILE.exists():
        with open(STATE_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return {
        "today": str(datetime.now().date()),
        "sent_today": 0,
        "last_sent": None,
        "total_sent": 0
    }


def save_state(state):
    """送信状態を保存"""
    with open(STATE_FILE, 'w', encoding='utf-8') as f:
        json.dump(state, f, ensure_ascii=False, indent=2)


def log_send(clinic_name, status, message=""):
    """送信ログを記録"""
    logs = []
    if LOG_FILE.exists():
        with open(LOG_FILE, 'r', encoding='utf-8') as f:
            logs = json.load(f)

    logs.append({
        "timestamp": datetime.now().isoformat(),
        "clinic": clinic_name,
        "status": status,
        "message": message
    })

    with open(LOG_FILE, 'w', encoding='utf-8') as f:
        json.dump(logs, f, ensure_ascii=False, indent=2)


def get_next_clinic():
    """次に送信する院を取得"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['自動送信可能']

    if ws.max_row < 2:
        wb.close()
        return None

    # 最初の院を取得
    clinic = {
        "row": 2,
        "name": ws.cell(row=2, column=2).value,
        "contact_url": ws.cell(row=2, column=10).value,
        "official_site": ws.cell(row=2, column=11).value,
        "message": ws.cell(row=2, column=13).value,
        "subject": ws.cell(row=2, column=14).value
    }

    wb.close()
    return clinic


def move_to_manual(clinic_name, reason):
    """院を手動リストに移動"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_auto = wb['自動送信可能']
    ws_manual = wb['手動送信のみ']

    # 該当する院を探す
    target_row = None
    for row in range(2, ws_auto.max_row + 1):
        if ws_auto.cell(row=row, column=2).value == clinic_name:
            target_row = row
            break

    if target_row:
        # データをコピー
        row_data = [ws_auto.cell(row=target_row, column=col).value for col in range(1, 15)]
        manual_last_row = ws_manual.max_row + 1

        for col, value in enumerate(row_data, start=1):
            ws_manual.cell(row=manual_last_row, column=col).value = value

        ws_manual.cell(row=manual_last_row, column=9).value = f"手動送信（{reason}）"

        # 自動リストから削除
        ws_auto.delete_rows(target_row)

        wb.save(EXCEL_FILE)
        logger.info(f"手動リストに移動: {clinic_name} - {reason}")

    wb.close()


def mark_as_sent(clinic_name):
    """送信済みとしてマーク（リストから削除）"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws_auto = wb['自動送信可能']

    for row in range(2, ws_auto.max_row + 1):
        if ws_auto.cell(row=row, column=2).value == clinic_name:
            ws_auto.delete_rows(row)
            break

    wb.save(EXCEL_FILE)
    wb.close()


def detect_form_fields(page):
    """フォームフィールドを検出"""
    fields = {
        "name": None,
        "furigana": None,
        "email": None,
        "tel": None,
        "message": None,
        "gender": None,
        "submit": None,
        "checkbox": None
    }

    # 名前フィールド
    for selector in ['input[name*="name"]', 'input[name*="氏名"]', 'input[name="your-name"]']:
        el = page.query_selector(selector)
        if el:
            fields["name"] = selector
            break

    # ふりがなフィールド
    for selector in ['input[name*="furigana"]', 'input[name*="kana"]', 'input[name="your-name2"]']:
        el = page.query_selector(selector)
        if el:
            fields["furigana"] = selector
            break

    # メールフィールド
    for selector in ['input[type="email"]', 'input[name*="mail"]', 'input[name*="email"]']:
        el = page.query_selector(selector)
        if el:
            fields["email"] = selector
            break

    # 電話フィールド
    for selector in ['input[name*="tel"]', 'input[name*="phone"]', 'input[type="tel"]']:
        el = page.query_selector(selector)
        if el:
            fields["tel"] = selector
            break

    # メッセージフィールド
    for selector in ['textarea[name*="message"]', 'textarea[name*="content"]', 'textarea[name*="inquiry"]', 'textarea']:
        el = page.query_selector(selector)
        if el:
            name = el.get_attribute('name')
            if name and 'recaptcha' not in name:
                fields["message"] = selector
                break

    # 送信ボタン
    for selector in ['input[type="submit"]', 'button[type="submit"]', 'button:has-text("送信")']:
        el = page.query_selector(selector)
        if el:
            fields["submit"] = selector
            break

    # チェックボックス（同意など）
    for selector in ['input[name*="acceptance"]', 'input[name*="agree"]', 'input[type="checkbox"]']:
        el = page.query_selector(selector)
        if el:
            fields["checkbox"] = selector

    return fields


def send_to_form(clinic):
    """フォームに送信"""
    SCREENSHOT_DIR.mkdir(exist_ok=True)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        page = browser.new_page()

        try:
            # フォームURLを決定
            url = clinic["contact_url"] or clinic["official_site"]
            if not url:
                return False, "URLなし"

            # 様々な問い合わせページURLパターンを試す
            base_url = url.rstrip('/')
            # ドメインのルートを取得
            from urllib.parse import urlparse
            parsed = urlparse(url)
            root_url = f"{parsed.scheme}://{parsed.netloc}"

            contact_urls = [
                url,  # 元のURL
                f"{base_url}/contact/",
                f"{base_url}/contact",
                f"{root_url}/contact/",
                f"{root_url}/contact",
                f"{root_url}/inquiry/",
                f"{root_url}/form/",
                f"{root_url}/mail/",
                f"{root_url}/otoiawase/",
                f"{root_url}/#contact",
            ]
            # 重複を除去
            contact_urls = list(dict.fromkeys(contact_urls))

            form_found = False
            for try_url in contact_urls:
                try:
                    page.goto(try_url, timeout=30000)
                    page.wait_for_load_state('networkidle', timeout=10000)

                    forms = page.query_selector_all('form')
                    if forms:
                        form_found = True
                        break
                except:
                    continue

            if not form_found:
                return False, "フォームなし"

            # フィールド検出
            fields = detect_form_fields(page)

            if not fields["message"]:
                return False, "メッセージ欄なし"

            # フォーム入力
            if fields["name"]:
                page.fill(fields["name"], SENDER_INFO["name"])

            if fields["furigana"]:
                page.fill(fields["furigana"], SENDER_INFO["name_kana"])

            if fields["email"]:
                page.fill(fields["email"], SENDER_INFO["email"])

            if fields["tel"]:
                page.fill(fields["tel"], SENDER_INFO["tel"])

            if fields["message"]:
                page.fill(fields["message"], clinic["message"])

            # すべてのチェックボックスをチェック（性別、同意など）
            all_checkboxes = page.query_selector_all('input[type="checkbox"]')
            for checkbox in all_checkboxes:
                try:
                    name = checkbox.get_attribute('name') or ''
                    # recaptcha以外のチェックボックスをチェック
                    if 'recaptcha' not in name.lower():
                        if not checkbox.is_checked():
                            checkbox.check()
                except:
                    pass

            # ラジオボタンも処理（性別など - 最初の選択肢を選ぶ）
            radio_groups = {}
            all_radios = page.query_selector_all('input[type="radio"]')
            for radio in all_radios:
                name = radio.get_attribute('name')
                if name and name not in radio_groups:
                    radio_groups[name] = radio
            for name, radio in radio_groups.items():
                try:
                    if not radio.is_checked():
                        radio.check()
                except:
                    pass

            # 少し待機してフォームが更新されるのを待つ
            time.sleep(1)

            # 送信前スクリーンショット
            screenshot_path = SCREENSHOT_DIR / f"{clinic['name']}_before.png"
            page.screenshot(path=str(screenshot_path))

            # 送信
            if fields["submit"]:
                # 送信ボタンが有効になるまで待機（最大10秒）
                submit_btn = page.query_selector(fields["submit"])
                for _ in range(20):
                    if submit_btn and submit_btn.is_enabled():
                        break
                    time.sleep(0.5)

                # それでもdisabledなら失敗
                if submit_btn and not submit_btn.is_enabled():
                    return False, "送信ボタン無効（必須項目不足）"

                page.click(fields["submit"])
                time.sleep(3)

                # 送信後スクリーンショット
                screenshot_path = SCREENSHOT_DIR / f"{clinic['name']}_after.png"
                page.screenshot(path=str(screenshot_path))

                # エラーチェック
                content = page.content()
                if 'エラー' in content or '入力内容に不備' in content or '必須' in content:
                    return False, "送信エラー"

                if '送信完了' in content or 'ありがとう' in content or '受け付けました' in content:
                    return True, "送信成功"

                # 成功かどうか判断が難しい場合
                return True, "送信完了（要確認）"
            else:
                return False, "送信ボタンなし"

        except Exception as e:
            return False, f"エラー: {str(e)}"
        finally:
            browser.close()


def run_once():
    """1回の送信サイクル"""
    state = load_state()

    # 日付が変わったらリセット
    today = str(datetime.now().date())
    if state["today"] != today:
        state["today"] = today
        state["sent_today"] = 0

    # 1日の上限チェック
    daily_limit = random.randint(DAILY_MIN, DAILY_MAX)
    if state["sent_today"] >= daily_limit:
        logger.info(f"本日の送信上限（{daily_limit}件）に達しました")
        return False

    # 次の院を取得
    clinic = get_next_clinic()
    if not clinic:
        logger.info("送信可能な院がありません。リスト補充が必要です。")
        return False

    logger.info(f"送信開始: {clinic['name']}")

    # 送信実行
    success, message = send_to_form(clinic)

    if success:
        mark_as_sent(clinic["name"])
        state["sent_today"] += 1
        state["total_sent"] += 1
        state["last_sent"] = datetime.now().isoformat()
        log_send(clinic["name"], "成功", message)
        logger.info(f"送信成功: {clinic['name']} - {message}")
    else:
        move_to_manual(clinic["name"], message)
        log_send(clinic["name"], "失敗", message)
        logger.warning(f"送信失敗: {clinic['name']} - {message}")

    save_state(state)
    return True


def run_daemon():
    """デーモンモードで実行（継続的に送信）"""
    logger.info("自動送信デーモン開始")

    while True:
        try:
            result = run_once()

            if result:
                # ランダム間隔で待機
                interval = random.randint(MIN_INTERVAL, MAX_INTERVAL)
                logger.info(f"次の送信まで {interval // 60} 分待機")
                time.sleep(interval)
            else:
                # 送信できない場合は1時間待機
                logger.info("1時間後に再試行")
                time.sleep(3600)

        except KeyboardInterrupt:
            logger.info("デーモン停止")
            break
        except Exception as e:
            logger.error(f"エラー: {e}")
            time.sleep(600)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="整骨院自動営業メール送信")
    parser.add_argument("--once", action="store_true", help="1回だけ実行")
    parser.add_argument("--daemon", action="store_true", help="デーモンモードで継続実行")
    parser.add_argument("--status", action="store_true", help="状態を表示")

    args = parser.parse_args()

    if args.status:
        state = load_state()
        print(f"本日: {state['today']}")
        print(f"本日の送信数: {state['sent_today']}")
        print(f"累計送信数: {state['total_sent']}")
        print(f"最終送信: {state['last_sent']}")
    elif args.once:
        run_once()
    elif args.daemon:
        run_daemon()
    else:
        parser.print_help()
