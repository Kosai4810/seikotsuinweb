#!/usr/bin/env python3
"""DNS認証完了を待ってメール送信を開始"""
import requests
import time
import json
import openpyxl
from datetime import datetime
from pathlib import Path

RESEND_API_KEY = "re_gyQhTCR9_M4nBeezWzVGB7kCWKqnVb15m"
DOMAIN_ID = "74b87383-a27d-42b2-81f1-9e7a7ce7ab26"
BASE_DIR = Path("/Users/kosai/Desktop/整骨院web/seikotsuin-web/営業")
EXCEL_FILE = BASE_DIR / "整骨院web_営業リスト_個別営業文付き.xlsx"
LOG_FILE = BASE_DIR / "送信ログ.json"

def check_verified():
    response = requests.get(
        f"https://api.resend.com/domains/{DOMAIN_ID}",
        headers={"Authorization": f"Bearer {RESEND_API_KEY}"}
    )
    return response.json().get('status') == 'verified'

def send_email(to, subject, body, clinic_name):
    response = requests.post(
        "https://api.resend.com/emails",
        headers={
            "Authorization": f"Bearer {RESEND_API_KEY}",
            "Content-Type": "application/json"
        },
        json={
            "from": "WEBBRIO <noreply@seikotsuinweb.com>",
            "to": [to],
            "subject": subject,
            "text": body,
            "reply_to": "contact@seikotsuinweb.com"
        }
    )
    return response.status_code == 200, response.json()

def log_send(clinic, status, message=""):
    logs = []
    if LOG_FILE.exists():
        with open(LOG_FILE, 'r', encoding='utf-8') as f:
            logs = json.load(f)
    logs.append({
        "timestamp": datetime.now().isoformat(),
        "clinic": clinic,
        "method": "メール（Resend）",
        "status": status,
        "message": message
    })
    with open(LOG_FILE, 'w', encoding='utf-8') as f:
        json.dump(logs, f, ensure_ascii=False, indent=2)

def mark_sent_in_excel(clinic_name):
    """Excelで送信済みマーク"""
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb['手動送信のみ']
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=2).value == clinic_name:
            ws.cell(row=row, column=9).value = f"送信済み（{datetime.now().strftime('%Y-%m-%d')}）"
            break
    wb.save(EXCEL_FILE)
    wb.close()

print("DNS認証待機中...")
while not check_verified():
    requests.post(
        f"https://api.resend.com/domains/{DOMAIN_ID}/verify",
        headers={"Authorization": f"Bearer {RESEND_API_KEY}"}
    )
    time.sleep(30)
    print(f"[{datetime.now().strftime('%H:%M:%S')}] まだpending...")

print("✓ DNS認証完了！")

# テスト送信
print("テストメール送信...")
success, result = send_email(
    "contact@seikotsuinweb.com",
    "【テスト】自動送信テスト",
    "Resend API自動送信テストです。",
    "テスト"
)
if success:
    print(f"✓ テスト成功: {result.get('id')}")
else:
    print(f"✗ テスト失敗: {result}")
    exit(1)

print("\n営業メール送信開始...")

# Excelから送信対象を取得
wb = openpyxl.load_workbook(EXCEL_FILE)
ws = wb['手動送信のみ']

sent_count = 0
for row in range(2, ws.max_row + 1):
    if sent_count >= 1:  # まず1件だけ
        break
    
    clinic_name = ws.cell(row=row, column=2).value
    contact_method = ws.cell(row=row, column=9).value or ""
    email_body = ws.cell(row=row, column=13).value
    subject = ws.cell(row=row, column=14).value
    
    # 送信済みスキップ
    if "送信済み" in contact_method:
        continue
    
    # メール本文があるもののみ
    if not email_body or not subject:
        continue
    
    # さくらリバース治療院の場合、メールアドレスを探す必要がある
    # とりあえず最初の1件を送信
    
    print(f"送信対象: {clinic_name}")
    # この院のメールアドレスを取得する必要あり
    # 今回はcontact@seikotsuinweb.comにテスト送信
    
    success, result = send_email(
        "contact@seikotsuinweb.com",  # テスト用
        subject.replace("〇〇", clinic_name.split("整骨院")[0] if "整骨院" in clinic_name else clinic_name[:5]),
        email_body,
        clinic_name
    )
    
    if success:
        print(f"✓ 送信成功: {clinic_name}")
        log_send(clinic_name, "成功", f"ID: {result.get('id')}")
        sent_count += 1
    else:
        print(f"✗ 送信失敗: {clinic_name} - {result}")
        log_send(clinic_name, "失敗", str(result))

wb.close()
print(f"\n送信完了: {sent_count}件")
